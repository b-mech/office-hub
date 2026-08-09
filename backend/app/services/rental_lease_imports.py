from __future__ import annotations

import asyncio
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
import json
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.rentals import RentalCompany, RentalLease, RentalLeaseImportBatch, RentalLeaseImportRow, RentalLeaseTenant, RentalProperty, RentalTenant, RentalUnit
from app.schemas.rentals import LeaseImportRowPatch, LeaseParsedData
from app.services.extraction.claude_provider import ClaudeProvider


HEADERS = ["Property Street Address", "Unit Label", "Tenant 1 Name", "Tenant 1 Phone", "Tenant 1 Email", "Tenant 2 Name", "Tenant 2 Phone", "Tenant 2 Email", "Tenant 3 Name", "Tenant 3 Phone", "Tenant 3 Email", "Rent", "Rent Discount", "Deposit", "Water Credit", "Lease Start", "Lease End", "Notes"]


async def generate_template(db: AsyncSession) -> bytes:
    units = (await db.execute(select(RentalUnit, RentalProperty).join(RentalProperty, RentalProperty.id == RentalUnit.property_id).order_by(RentalProperty.street_address, RentalUnit.unit_label))).all()
    workbook = Workbook()
    rent_roll = workbook.active
    rent_roll.title = "Rent Roll"
    rent_roll.append(HEADERS)
    for cell in rent_roll[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A527A")
    for _ in range(100):
        rent_roll.append([None] * len(HEADERS))
    for row in range(2, 102):
        rent_roll.cell(row, 16).number_format = "yyyy-mm-dd"
        rent_roll.cell(row, 17).number_format = "yyyy-mm-dd"
    rent_roll.freeze_panes = "A2"
    rent_roll.auto_filter.ref = f"A1:R101"
    for column in rent_roll.columns:
        rent_roll.column_dimensions[column[0].column_letter].width = min(max(len(str(column[0].value or "")) + 3, 14), 28)

    reference = workbook.create_sheet("Existing Units")
    reference.append(["Street Address", "Unit Label", "Current Tenant(s)", "Current Rent", "Lease End"])
    for cell in reference[1]:
        cell.font = Font(bold=True)
    for unit, rental_property in units:
        lease = await db.scalar(select(RentalLease).where(RentalLease.unit_id == unit.id, RentalLease.status.in_(["active", "month_to_month"])).order_by(RentalLease.lease_start.desc().nullslast()).limit(1))
        tenants: list[str] = []
        if lease:
            tenants = list((await db.scalars(select(RentalTenant.full_name).join(RentalLeaseTenant, RentalLeaseTenant.tenant_id == RentalTenant.id).where(RentalLeaseTenant.lease_id == lease.id))).all())
        reference.append([rental_property.street_address, unit.unit_label, ", ".join(tenants), lease.rent if lease else None, lease.lease_end if lease else None])
    reference.protection.sheet = True
    for row in reference.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    reference.freeze_panes = "A2"

    instructions = workbook.create_sheet("Instructions")
    lines = [
        "Lease Import Instructions",
        "Enter one lease per row on Rent Roll. Do not change the header names.",
        "Copy Street Address and Unit Label from Existing Units whenever possible.",
        "For a new property or unit, type the new address/unit directly.",
        "Use real Excel date cells for Lease Start and Lease End (YYYY-MM-DD display).",
        "For a 4th+ tenant add: Additional tenant: Name / Phone / Email in Notes.",
        "Every uploaded row must be reviewed and approved before live lease data changes.",
    ]
    for line in lines:
        instructions.append([line])
    instructions["A1"].font = Font(bold=True, size=16)
    instructions.column_dimensions["A"].width = 100
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _json_cell(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


async def upload_and_extract(db: AsyncSession, filename: str, content: bytes) -> RentalLeaseImportBatch:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if "Rent Roll" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a 'Rent Roll' sheet")
    sheet = workbook["Rent Roll"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if headers[: len(HEADERS)] != HEADERS:
        raise ValueError("Rent Roll headers do not match the Office Hub lease-import template")
    batch = RentalLeaseImportBatch(source_filename=filename, status="processing")
    db.add(batch)
    await db.flush()
    rows: list[RentalLeaseImportRow] = []
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True), start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        raw = {header: _json_cell(value) for header, value in zip(HEADERS, values, strict=True)}
        row = RentalLeaseImportRow(batch_id=batch.id, source_row_number=source_row, raw_data=raw)
        db.add(row)
        rows.append(row)
    batch.total_rows = len(rows)
    batch.rows_pending = len(rows)
    await db.flush()
    context = await _matching_context(db)
    valid_unit_ids = {unit["id"] for unit in context["units"]}
    valid_lease_ids = {unit["active_lease"]["id"] for unit in context["units"] if unit["active_lease"]}
    for row in rows:
        try:
            result = await asyncio.to_thread(_extract_row, row.raw_data, context)
            parsed = LeaseParsedData.model_validate(result["parsed_data"])
            row.parsed_data = parsed.model_dump(mode="json")
            row.confidence = result.get("confidence") or {"flags": ["Claude returned no confidence values"]}
            flags = row.confidence.setdefault("flags", [])
            row.match_type = result.get("match_type") if result.get("match_type") in {"existing_unit", "new_unit", "unresolved"} else "unresolved"
            matched_unit_id = result.get("matched_unit_id")
            row.matched_unit_id = matched_unit_id if matched_unit_id in valid_unit_ids else None
            if matched_unit_id is not None and row.matched_unit_id is None:
                row.match_type = "unresolved"
                flags.append(f"Claude returned unknown unit id {matched_unit_id}")
            row.suggested_action = result.get("suggested_action") if result.get("suggested_action") in {"create_lease", "renew_lease", "update_lease", "skip"} else "skip"
            existing_lease_id = result.get("existing_lease_id")
            row.existing_lease_id = existing_lease_id if existing_lease_id in valid_lease_ids else None
            if row.suggested_action in {"renew_lease", "update_lease"} and row.existing_lease_id is None:
                flags.append("Renew/update requires a confirmed existing active lease")
        except Exception as exc:
            row.parsed_data = None
            row.confidence = {"flags": [f"Parse failed; manual entry required: {type(exc).__name__}: {exc}"]}
            row.match_type = "unresolved"
            row.suggested_action = "skip"
    batch.status = "needs_review"
    await db.commit()
    await db.refresh(batch)
    return batch


async def _matching_context(db: AsyncSession) -> dict[str, Any]:
    rows = (await db.execute(select(RentalUnit, RentalProperty).join(RentalProperty))).all()
    units = []
    for unit, rental_property in rows:
        lease = await db.scalar(select(RentalLease).where(RentalLease.unit_id == unit.id, RentalLease.status.in_(["active", "month_to_month"])).order_by(RentalLease.lease_start.desc().nullslast()).limit(1))
        tenant_names: list[str] = []
        if lease:
            tenant_names = list((await db.scalars(select(RentalTenant.full_name).join(RentalLeaseTenant, RentalLeaseTenant.tenant_id == RentalTenant.id).where(RentalLeaseTenant.lease_id == lease.id))).all())
        units.append({"id": unit.id, "street_address": rental_property.street_address, "unit_label": unit.unit_label, "active_lease": None if not lease else {"id": lease.id, "tenant_names": tenant_names, "rent": str(lease.rent), "lease_start": lease.lease_start.isoformat() if lease.lease_start else None, "lease_end": lease.lease_end.isoformat() if lease.lease_end else None}})
    return {"units": units}


def _extract_row(raw_data: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    provider = ClaudeProvider()
    system = """You extract one rental lease spreadsheet row. Return strict JSON only with keys parsed_data, match_type, matched_unit_id, suggested_action, existing_lease_id, confidence. parsed_data must contain property_street_address, unit_label, tenants, rent, rent_discount_amount, deposit, water_credit_amount, lease_start, lease_end, lease_notes. Each tenant has full_name, phone, email, is_primary_contact. Extract additional tenants from Notes. confidence contains 0-1 values per parsed field and flags, an array of uncertainty strings. Match only against the supplied units. Use existing_unit, new_unit, or unresolved. Use create_lease when no active lease exists; renew_lease for overlapping tenants with a new term; update_lease for same tenants and same start date with corrections; skip for junk."""
    response = provider.client.messages.create(model=provider.model_version, max_tokens=3000, system=system, messages=[{"role": "user", "content": json.dumps({"raw_row": raw_data, "current_rentals": context}, default=str)}])
    raw = provider._extract_text_response(response)
    return provider._parse_json_response(raw)


async def patch_row(db: AsyncSession, row: RentalLeaseImportRow, patch: LeaseImportRowPatch) -> RentalLeaseImportRow:
    changes = patch.model_dump(exclude_unset=True)
    if "parsed_data" in changes and changes["parsed_data"] is not None:
        changes["parsed_data"] = patch.parsed_data.model_dump(mode="json") if patch.parsed_data else None
    for key, value in changes.items():
        setattr(row, key, value)
    row.review_status = "edited"
    await _refresh_batch(db, row.batch_id)
    await db.commit()
    await db.refresh(row)
    return row


async def approve_row(db: AsyncSession, row: RentalLeaseImportRow) -> RentalLeaseImportRow:
    if row.review_status in {"approved", "rejected"}:
        raise ValueError("This import row has already been reviewed")
    if row.suggested_action == "skip":
        row.review_status = "approved"
        row.reviewed_at = func.now()
    else:
        if not row.parsed_data:
            raise ValueError("Parsed lease data is required before approval")
        parsed = LeaseParsedData.model_validate(row.parsed_data)
        unit_id = row.matched_unit_id
        if row.match_type == "new_unit":
            unit_id = await _create_or_match_unit(db, parsed)
            row.matched_unit_id = unit_id
        if not unit_id or row.match_type == "unresolved":
            raise ValueError("A resolved rental unit is required before approval")
        if row.suggested_action == "update_lease":
            if not row.existing_lease_id:
                raise ValueError("An existing lease is required for update")
            lease = await db.get(RentalLease, row.existing_lease_id)
            if not lease:
                raise ValueError("Existing lease was not found")
            existing_notes = lease.lease_notes
            before = {"rent": str(lease.rent), "deposit": str(lease.deposit), "lease_start": str(lease.lease_start), "lease_end": str(lease.lease_end)}
            _apply_lease_values(lease, parsed)
            change_note = f"[lease import update] Previous values: {json.dumps(before)}"
            lease.lease_notes = "\n".join(filter(None, [existing_notes, change_note, parsed.lease_notes]))
        else:
            if row.suggested_action not in {"create_lease", "renew_lease"}:
                raise ValueError("Select a valid lease action before approval")
            lease = RentalLease(unit_id=unit_id, rent=parsed.rent)
            _apply_lease_values(lease, parsed)
            db.add(lease)
            await db.flush()
            if row.suggested_action == "renew_lease" and row.existing_lease_id:
                prior = await db.get(RentalLease, row.existing_lease_id)
                if prior:
                    prior.status = "expired"
                    proposed_end = parsed.lease_start - timedelta(days=1) if parsed.lease_start else prior.lease_end
                    if proposed_end and (prior.lease_end is None or prior.lease_end < proposed_end):
                        prior.lease_end = proposed_end
        await _replace_tenants(db, lease, parsed)
        row.committed_lease_id = lease.id
        row.review_status = "approved"
        row.reviewed_at = func.now()
    await _refresh_batch(db, row.batch_id)
    await db.commit()
    await db.refresh(row)
    return row


def _apply_lease_values(lease: RentalLease, parsed: LeaseParsedData) -> None:
    lease.rent = parsed.rent
    lease.rent_discount_amount = parsed.rent_discount_amount
    lease.deposit = parsed.deposit
    lease.lease_start = parsed.lease_start
    lease.lease_end = parsed.lease_end
    lease.status = "expired" if parsed.lease_end and parsed.lease_end < date.today() else "active"
    if parsed.lease_notes:
        lease.lease_notes = parsed.lease_notes


async def _create_or_match_unit(db: AsyncSession, parsed: LeaseParsedData) -> int:
    company_id = await db.scalar(select(RentalCompany.id).where(func.lower(RentalCompany.name) == "privi"))
    if not company_id:
        raise ValueError("PRIVI rental company is missing")
    rental_property = await db.scalar(select(RentalProperty).where(RentalProperty.company_id == company_id, func.lower(RentalProperty.street_address) == parsed.property_street_address.casefold()))
    if not rental_property:
        rental_property = RentalProperty(company_id=company_id, street_address=parsed.property_street_address, property_type="residential")
        db.add(rental_property)
        await db.flush()
    unit = await db.scalar(select(RentalUnit).where(RentalUnit.property_id == rental_property.id, RentalUnit.unit_label.is_not_distinct_from(parsed.unit_label)))
    if not unit:
        unit = RentalUnit(property_id=rental_property.id, unit_label=parsed.unit_label)
        db.add(unit)
        await db.flush()
    return unit.id


async def _replace_tenants(db: AsyncSession, lease: RentalLease, parsed: LeaseParsedData) -> None:
    await db.execute(RentalLeaseTenant.__table__.delete().where(RentalLeaseTenant.lease_id == lease.id))
    for position, incoming in enumerate(parsed.tenants):
        tenant = await db.scalar(select(RentalTenant).where(func.lower(RentalTenant.full_name) == incoming.full_name.casefold()))
        if not tenant:
            tenant = RentalTenant(full_name=incoming.full_name, phone=incoming.phone, email=incoming.email)
            db.add(tenant)
            await db.flush()
        else:
            tenant.phone = incoming.phone or tenant.phone
            tenant.email = incoming.email or tenant.email
        db.add(RentalLeaseTenant(lease_id=lease.id, tenant_id=tenant.id, is_primary_contact=incoming.is_primary_contact or position == 0))


async def reject_row(db: AsyncSession, row: RentalLeaseImportRow) -> RentalLeaseImportRow:
    if row.review_status in {"approved", "rejected"}:
        raise ValueError("This import row has already been reviewed")
    row.review_status = "rejected"
    row.reviewed_at = func.now()
    await _refresh_batch(db, row.batch_id)
    await db.commit()
    await db.refresh(row)
    return row


async def approve_all_clean(db: AsyncSession, batch_id: int) -> tuple[int, int]:
    rows = list((await db.scalars(select(RentalLeaseImportRow).where(RentalLeaseImportRow.batch_id == batch_id, RentalLeaseImportRow.review_status.in_(["needs_review", "edited"])))).all())
    approved = 0
    skipped = 0
    for row in rows:
        flags = (row.confidence or {}).get("flags") or []
        if flags or row.match_type == "unresolved":
            skipped += 1
            continue
        await approve_row(db, row)
        approved += 1
    return approved, skipped


async def _refresh_batch(db: AsyncSession, batch_id: int) -> None:
    await db.flush()
    batch = await db.get(RentalLeaseImportBatch, batch_id)
    if not batch:
        return
    pending = await db.scalar(select(func.count()).select_from(RentalLeaseImportRow).where(RentalLeaseImportRow.batch_id == batch_id, RentalLeaseImportRow.review_status.in_(["needs_review", "edited"])))
    batch.rows_pending = int(pending or 0)
    batch.status = "closed" if batch.rows_pending == 0 else "needs_review"
