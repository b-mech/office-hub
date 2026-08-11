from __future__ import annotations

import argparse
import asyncio
from dataclasses import dataclass
from datetime import datetime
from datetime import timezone
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select

from app.modules.costbook.models import PurchaseOrder  # noqa: F401 - registers ORM relationship target
from app.core.database import AsyncSessionLocal
from app.models.core import Development
from app.models.core import LegalDescriptionVerificationStatus
from app.models.core import Lot
from app.models.core import LotStatus
from app.models.core import LotTriggerType
from app.models.financing import Property
from app.services.developments import DevelopmentService


REQUIRED_COLUMNS = {
    "civic_address",
    "property_id",
    "block",
    "lot_number",
    "plan",
    "legal_confirmed",
    "legal_description_confidence",
    "source_document",
    "development_name",
    "development_path",
    "trigger_type",
    "lifecycle_status",
}
VALID_TRIGGERS = {"otp", "spec", "showhome"}
VALID_STATUSES = {
    "land_contracted", "land_purchased", "serviced", "sale_signed",
    "build_active", "possession", "warranty",
}
VERIFICATION_STATUS_MAP = {
    "title-confirmed": LegalDescriptionVerificationStatus.TITLE_CONFIRMED,
    "title_confirmed": LegalDescriptionVerificationStatus.TITLE_CONFIRMED,
    "permit/agreement-confirmed": LegalDescriptionVerificationStatus.PERMIT_AGREEMENT_CONFIRMED,
    "permit_agreement_confirmed": LegalDescriptionVerificationStatus.PERMIT_AGREEMENT_CONFIRMED,
}


@dataclass(frozen=True)
class ImportRow:
    row_number: int
    values: dict[str, str]

    @property
    def normalized_legal(self) -> str:
        return f"BLK {self.values['block']} LT {self.values['lot_number']} PLAN {self.values['plan']}"


def read_rows(path: Path) -> list[ImportRow]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["Lots to complete"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    missing = REQUIRED_COLUMNS.difference(headers)
    if missing:
        raise ValueError(f"Workbook is missing columns: {', '.join(sorted(missing))}")
    rows: list[ImportRow] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {headers[index]: str(value).strip() if value is not None else "" for index, value in enumerate(cells)}
        if values.get("civic_address"):
            rows.append(ImportRow(row_number=row_number, values=values))
    return rows


async def validate(path: Path, *, apply: bool = False) -> bool:
    rows = read_rows(path)
    ready = True
    seen_legals: set[str] = set()
    async with AsyncSessionLocal() as db:
        developments = list((await db.scalars(select(Development))).all())
        development_service = DevelopmentService(db)
        paths_by_id: dict[UUID, str] = {}
        for org_id in {development.org_id for development in developments}:
            for development, full_path in await development_service.list_with_paths(org_id):
                paths_by_id[development.id] = full_path
        paths_by_key = {
            paths_by_id[development.id].casefold(): development
            for development in developments
        }
        pending: list[tuple[ImportRow, Property, Development, LegalDescriptionVerificationStatus]] = []
        for row in rows:
            errors: list[str] = []
            for field in (
                "property_id", "block", "lot_number", "plan", "development_name", "development_path",
                "trigger_type", "lifecycle_status", "legal_description_confidence", "source_document",
            ):
                if not row.values.get(field):
                    errors.append(f"missing {field}")
            if row.values.get("legal_confirmed").casefold() != "yes":
                errors.append("legal description is not confirmed")
            if row.values.get("trigger_type") and row.values["trigger_type"].casefold() not in VALID_TRIGGERS:
                errors.append("invalid trigger_type")
            if row.values.get("lifecycle_status") and row.values["lifecycle_status"].casefold() not in VALID_STATUSES:
                errors.append("invalid lifecycle_status")
            verification_status = VERIFICATION_STATUS_MAP.get(row.values.get("legal_description_confidence", "").casefold())
            if verification_status is None:
                errors.append("invalid legal_description_confidence")
            try:
                property_id = UUID(row.values["property_id"])
            except (ValueError, KeyError):
                property_id = None
                errors.append("invalid property_id")
            property_record = await db.get(Property, property_id) if property_id is not None else None
            if property_id is not None and property_record is None:
                errors.append("property_id does not exist")
            development_path = row.values.get("development_path", "")
            development = paths_by_key.get(development_path.casefold()) if development_path else None
            if development is None:
                errors.append("development_path does not resolve to one record")
            elif development.name.casefold() != row.values.get("development_name", "").casefold():
                errors.append("development_name does not match the development_path leaf")
            if row.values.get("block") and row.values.get("lot_number") and row.values.get("plan"):
                legal = row.normalized_legal
                if legal in seen_legals:
                    errors.append("duplicate legal description within workbook")
                seen_legals.add(legal)
                if await db.scalar(select(Lot.id).where(Lot.legal_description_normalized == legal)):
                    errors.append("legal description already exists in core.lots")
            status = "READY" if not errors else "INCOMPLETE"
            print(f"Row {row.row_number} — {row.values['civic_address']}: {status}")
            for error in errors:
                print(f"  - {error}")
            ready = ready and not errors
            if not errors and property_record is not None and development is not None and verification_status is not None:
                pending.append((row, property_record, development, verification_status))

        if apply and ready:
            verified_at = datetime.now(timezone.utc)
            for row, property_record, development, verification_status in pending:
                lot = Lot(
                    development_id=development.id,
                    property_id=property_record.id,
                    trigger_type=LotTriggerType(row.values["trigger_type"].casefold()),
                    legal_description_raw=row.values.get("legal_description_raw") or row.normalized_legal,
                    legal_description_normalized=row.normalized_legal,
                    civic_address=row.values["civic_address"],
                    lot_number=row.values["lot_number"],
                    block=row.values["block"],
                    plan=row.values["plan"],
                    status=LotStatus(row.values["lifecycle_status"].casefold()),
                    legal_description_verification_status=verification_status,
                    legal_description_source=row.values["source_document"],
                    legal_description_verified_at=verified_at,
                )
                db.add(lot)
            await db.commit()
            print(f"\nApply complete: {len(pending)} lots created in one transaction.")
        else:
            print(f"\nDry run complete: {len(rows)} rows checked; {'all rows ready' if ready else 'no data written'}.")
    return ready


async def main() -> None:
    parser = argparse.ArgumentParser(description="Validate the reviewed Office Hub lot import workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--apply", action="store_true", help="Create all validated lots in one transaction.")
    args = parser.parse_args()
    ready = await validate(args.workbook, apply=args.apply)
    if not ready:
        raise SystemExit(2)


if __name__ == "__main__":
    asyncio.run(main())
