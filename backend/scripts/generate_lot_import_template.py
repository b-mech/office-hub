from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_PATH = Path(__file__).resolve().parents[2] / "docs" / "imports" / "lender-allocation-lots-prefill.xlsx"

HEADERS = [
    "civic_address",
    "property_id",
    "existing_property_address",
    "block",
    "lot_number",
    "plan",
    "legal_description_raw",
    "legal_description_normalized",
    "legal_confirmed",
    "legal_description_confidence",
    "source_document",
    "development_name",
    "development_path",
    "development_confirmation_required",
    "municipality",
    "trigger_type",
    "lifecycle_status",
    "lifecycle_status_confirmation_required",
    "construction_stage",
    "notes_for_review",
]

ROWS = [
    [
        "22 Oak Meadow Drive",
        "afe24519-8e9d-47b9-a2cc-85c92f87b5f1",
        "22/24 Oak Meadow Drive – SPEC",
        "2", "17", "70771", "LOT 17 BLOCK 2 PLAN 70771 WLTO IN NW 1/4 21-11-5 EPM",
        "BLK 2 LT 17 PLAN 70771", "yes", "title-confirmed",
        "22 Oak Meadow Drive - Legal Docs Purchase.pdf", "Oakbank", "RM of Springfield → Oakbank", "no",
        "RM of Springfield", "spec", "build_active", "no", "LOCKUP",
        "Title 3327587/1. Development name is inferred from the Box community folder. Shares one property_id with 24 Oak Meadow; confirm that link.",
    ],
    [
        "24 Oak Meadow Drive",
        "afe24519-8e9d-47b9-a2cc-85c92f87b5f1",
        "22/24 Oak Meadow Drive – SPEC",
        "2", "18", "70771", "LOT 18 BLOCK 2 PLAN 70771 WLTO IN NW 1/4 21-11-5 EPM",
        "BLK 2 LT 18 PLAN 70771", "yes", "title-confirmed",
        "24 Oak Meadow Drive - Legal Docs Purchase.pdf", "Oakbank", "RM of Springfield → Oakbank", "no",
        "RM of Springfield", "spec", "build_active", "no", "LOCKUP",
        "Title 3327588/1. Development name is inferred from the Box community folder. Shares one property_id with 22 Oak Meadow; confirm that link.",
    ],
    [
        "27 Morning Glory Way",
        "45671ed0-d126-48e8-b438-88d82ed4ba44",
        "27 Morning Glory Way - SPEC",
        "4", "46", "74032",
        "LOT 46 BLOCK 4 PLAN 74032 WLTO IN NW 1/4 34-10-5 EPM",
        "BLK 4 LT 46 PLAN 74032",
        "yes", "title-confirmed", "27 Morning Glory Way - Title.pdf",
        "Dugald", "RM of Springfield → Dugald", "no", "RM of Springfield", "spec", "land_purchased", "no", "NA",
        "Title 3323554/1. Dugald is inferred from the Box production folder. land_purchased is inferred because no build start or active construction stage is recorded.",
    ],
    [
        "245 Blossom Way",
        "d2adf054-3c61-4060-84a5-49c6ae4cea08",
        "245 Blossom Way – SPEC",
        "4", "11", "71499", "LOT 11 BLOCK 4 PLAN 71499", "BLK 4 LT 11 PLAN 71499",
        "yes", "permit/agreement-confirmed", "1Q - 245 Blossom Way - Approved Permit RRPD-2026-685.pdf",
        "PARKVIEW POINTE", "RM of West St. Paul → PARKVIEW POINTE", "no", "RM of West St. Paul", "spec", "build_active", "no", "NA",
        "Lot 11 is independently supported by Property Assessment Lot Info.png. Parkview is supported by the Phase 3 OTP filename and Box path. Build start is 2026-07-27.",
    ],
    [
        "256 Middlechurch Gate",
        "428cacdd-32f0-4aec-828e-aa3b483786ac",
        "256 Middle Church",
        "1", "19", "71499", "LOT 19 BLOCK 1 PLAN 71499 WLTO", "BLK 1 LT 19 PLAN 71499",
        "yes", "title-confirmed", "1D - 256 Middlechurch Gate - Status of Title.pdf",
        "PARKVIEW POINTE", "RM of West St. Paul → PARKVIEW POINTE", "no", "RM of West St. Paul", "spec", "land_purchased", "no", "NA",
        "Middlechurch is confirmed as part of Parkview Pointe by its Box path and the Parkview Pointe Phase 3 bulk OTP filename. No build start/stage is recorded.",
    ],
    [
        "187 Middlechurch Gate",
        "f543d114-73d3-4263-a290-fc17ded920d5",
        "187 Middle Church – SPEC",
        "2", "2", "71499", "LOT 2 BLOCK 2 PLAN 71499", "BLK 2 LT 2 PLAN 71499",
        "yes", "permit/agreement-confirmed", "1Q - 187 Middlechurch Gate - Approved Permit RRPD-2026-680.pdf",
        "PARKVIEW POINTE", "RM of West St. Paul → PARKVIEW POINTE", "no", "RM of West St. Paul", "spec", "land_purchased", "no", "NA",
        "Middlechurch is confirmed as part of Parkview Pointe by its Box path and the Parkview Pointe Phase 3 bulk OTP filename. No build start/stage was found.",
    ],
    [
        "87 Grove Crescent",
        "e5a5034a-5a42-4035-8e33-05d3a64f6ab3",
        "87 Grove Crescent – SHOWHOME",
        "2", "13", "75588", "LOT 13 BLOCK 2 PLAN 75588 WLTO", "BLK 2 LT 13 PLAN 75588",
        "yes", "title-confirmed", "1D - 87 Grove Cres - Status of Title.pdf",
        "Forest Grove Estates", "Headingley → Forest Grove Estates", "no", "RM of Headingley", "showhome", "build_active", "no", "FOUNDATION",
        "The Forest Grove Phase 2 agreement schedule also maps this address to Block 2 Lot 13 Plan 75588. Construction sync says SPEC; property naming identifies SHOWHOME.",
    ],
    [
        "14 Grove Crescent",
        "60991f41-940e-4c5d-a83c-8232f41194a0",
        "14 Grove Cresent (Fall 2027)",
        "1", "4", "75588", "LOT 4 BLOCK 1 PLAN 75588", "BLK 1 LT 4 PLAN 75588",
        "yes", "permit/agreement-confirmed", "Forest Grove Estates Phase 2 - OTP Purchase.pdf",
        "Forest Grove Estates", "Headingley → Forest Grove Estates", "no", "RM of Headingley", "showhome", "land_purchased", "no", "NA",
        "The agreement schedule maps this address to Block 1 Lot 4 Plan 75588. Existing property address misspells Crescent. Fall 2027/no stage supports land_purchased as the provisional status.",
    ],
]


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lots to complete"
    sheet.append(HEADERS)
    for row in ROWS:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="17698C")
    attention_fill = PatternFill("solid", fgColor="FEF3C7")
    confirmed_fill = PatternFill("solid", fgColor="DCFCE7")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_number in range(2, sheet.max_row + 1):
        for cell in sheet[row_number]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        confirmation = sheet.cell(row=row_number, column=HEADERS.index("legal_confirmed") + 1)
        confirmation.fill = confirmed_fill if confirmation.value == "yes" else attention_fill
        for field in ("block", "lot_number", "plan", "development_name", "lifecycle_status"):
            cell = sheet.cell(row=row_number, column=HEADERS.index(field) + 1)
            if cell.value in (None, ""):
                cell.fill = attention_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 28, "B": 38, "C": 40, "D": 10, "E": 12, "F": 12, "G": 48,
        "H": 32, "I": 16, "J": 28, "K": 52, "L": 28, "M": 42, "N": 18,
        "O": 22, "P": 14, "Q": 18, "R": 20, "S": 20, "T": 70,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    yes_no = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    trigger = DataValidation(type="list", formula1='"otp,spec,showhome"', allow_blank=True)
    status = DataValidation(
        type="list",
        formula1='"land_contracted,land_purchased,serviced,sale_signed,build_active,possession,warranty"',
        allow_blank=True,
    )
    sheet.add_data_validation(yes_no)
    sheet.add_data_validation(trigger)
    sheet.add_data_validation(status)
    yes_no.add(f"I2:I{sheet.max_row}")
    yes_no.add(f"N2:N{sheet.max_row}")
    yes_no.add(f"R2:R{sheet.max_row}")
    trigger.add(f"P2:P{sheet.max_row}")
    status.add(f"Q2:Q{sheet.max_row}")

    instructions = workbook.create_sheet("Instructions")
    instructions.column_dimensions["A"].width = 115
    instructions["A1"] = "Office Hub lot import — completion instructions"
    instructions["A1"].font = Font(size=16, bold=True)
    instructions["A3"] = "Yellow cells require review or completion. Do not change property_id values unless the property relationship itself is being corrected."
    instructions["A4"] = "All eight legal descriptions are prefilled. legal_description_confidence distinguishes registered-title evidence from permit/developer-agreement evidence."
    instructions["A5"] = "Review every development_name and lifecycle_status marked confirmation required. Change the value if needed, then set its confirmation-required cell to no."
    instructions["A6"] = "22 and 24 Oak Meadow currently share one combined property record. State in notes whether both lots should retain that same property_id."
    instructions["A7"] = "Return the completed workbook unchanged. The importer runs in dry-run mode first and will not create lots until the validation report is reviewed."
    for row in instructions.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
