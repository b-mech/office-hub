from __future__ import annotations

import argparse
import asyncio
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import AsyncSessionLocal


UNIT_MAPPINGS = [
    ("Guay", "59 Guay Avenue", None, "Unit 1", False, 1, "1-59 Guay"),
    ("Guay", "59 Guay Avenue", None, "Unit 2", False, 2, "2-59 Guay"),
    ("Guay", "59 Guay Avenue", None, "Unit 3", True, 3, "3-59 Guay Base."),
    ("Guay", "59 Guay Avenue", None, "Unit 4", True, 4, "4-59 Guay Base."),
    ("Regent", "311 Regent Avenue", None, None, False, 7, "311 Regent"),
    ("Regent", "311 Regent Avenue", None, "Unit A", True, 8, "311-A Regent Base."),
    ("Regent", "313 Regent Avenue", None, None, False, 9, "313 Regent"),
    ("Regent", "313 Regent Avenue", None, "Unit A", True, 10, "313-A Regent Base."),
    ("Regent", "315 Regent Avenue", None, None, False, 11, "315 Regent"),
    ("Regent", "315 Regent Avenue", None, "Unit A", True, 12, "315-A Regent Base."),
    ("Lorette", "142 Clubhouse Drive", None, None, False, 13, "142 Clubhouse"),
    ("Lorette", "146 Clubhouse Drive", None, None, False, 14, "146 Clubhouse"),
    ("Lorette", "150 Clubhouse Drive", None, None, False, 15, "150 Clubhouse"),
    ("Lorette", "154 Clubhouse Drive", None, None, False, 16, "154 Clubhouse"),
    ("Lorette", "158 Clubhouse Drive", None, None, False, 17, "158 Clubhouse"),
    ("Lorette", "162 Clubhouse Drive", None, None, False, 18, "162 Clubhouse"),
    ("Lorette", "166 Clubhouse Drive", None, None, False, 19, "166 Clubhouse"),
    ("Lorette", "170 Clubhouse Drive", None, None, False, 20, "170 Clubhouse"),
    ("Landmark", "61 Sand Piper", "57 A Sand Piper", None, False, 21, "57A Sand Piper,now 61 Sand Piper"),
    ("Landmark", "57 Sand Piper", "57 B Sand Piper", None, False, 22, "57B Sand Piper,now 57 Sand Piper"),
    ("Landmark", "69 Sand Piper", "65 A Sand Piper", None, False, 23, "65A Sand Piper,now 69 Sand Piper"),
    ("Landmark", "65 Sand Piper", "65 B Sand Piper", None, False, 24, "65B Sand Piper,now 65 Sand Piper"),
    ("Grande Pointe", "79-100 GPM condo", None, None, False, 26, "79-100 GMP condo"),
    ("PRIVI - Commercial", "90 Froese Crescent", None, None, False, None, "90 Froese Crescent"),
]

RENT_COLUMNS = {
    "number": 0,
    "rent": 7,
    "water": 8,
    "discount": 9,
    "deposit": 10,
    "lease_start": 11,
    "lease_end": 12,
    "notes": 13,
}


def decimal_or_none(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, Decimal)):
        return Decimal(value)
    cleaned = str(value).strip().replace("$", "").replace(",", "")
    if not cleaned or set(cleaned) <= {"-", " "}:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def parse_discount(value: Any) -> tuple[Decimal | None, str | None]:
    amount = decimal_or_none(value)
    if amount is not None:
        return amount, None
    raw = "" if value is None else str(value).strip()
    if not raw or not raw.replace("$", "").replace("-", "").strip():
        return None, None
    return None, raw


def parse_water(value: Any) -> tuple[Decimal | None, str | None]:
    if value is None or str(value).strip() == "":
        return None, None
    raw = str(value).strip()
    match = re.search(r"\d+(?:\.\d+)?", raw)
    return (Decimal(match.group()) if match else None), raw


def parse_date(value: Any) -> tuple[date | None, str | None]:
    if value is None or str(value).strip() == "":
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    raw = str(value).strip()
    for fmt in ("%b-%d-%y", "%B %d,%Y", "%B %d, %Y", "%Y-%m-%d", "%m-%d-%y"):
        try:
            return datetime.strptime(raw, fmt).date(), None
        except ValueError:
            continue
    return None, raw


def score_and_notes(value: Any) -> tuple[int | None, str | None]:
    if value is None or str(value).strip() == "":
        return None, None
    raw = str(value).strip()
    score_text, separator, notes = raw.partition("/")
    try:
        score = int(score_text.strip())
    except ValueError:
        return None, raw
    return score, notes.strip() if separator and notes.strip() else None


def append_notes(parts: list[str], value: Any) -> None:
    if value is not None and str(value).strip():
        parts.append(str(value).strip())


async def get_or_create_company(db: AsyncSession) -> tuple[int, bool]:
    existing = await db.scalar(text("SELECT id FROM rental_companies WHERE lower(name) = 'privi'"))
    if existing is not None:
        return int(existing), False
    company_id = await db.scalar(text("INSERT INTO rental_companies (name) VALUES ('PRIVI') RETURNING id"))
    return int(company_id), True


async def upsert_property(db: AsyncSession, company_id: int, mapping: tuple[Any, ...]) -> tuple[int, bool]:
    group_name, address, former_address, *_ = mapping
    existing = await db.scalar(
        text("SELECT id FROM rental_properties WHERE company_id = :company_id AND street_address = :address"),
        {"company_id": company_id, "address": address},
    )
    property_type = "commercial" if group_name == "PRIVI - Commercial" else "residential"
    if existing is not None:
        await db.execute(
            text("UPDATE rental_properties SET group_name=:group_name, former_address=:former_address, property_type=:property_type, updated_at=now() WHERE id=:id"),
            {"id": existing, "group_name": group_name, "former_address": former_address, "property_type": property_type},
        )
        return int(existing), False
    property_id = await db.scalar(
        text("INSERT INTO rental_properties (company_id, group_name, street_address, former_address, property_type) VALUES (:company_id, :group_name, :address, :former_address, :property_type) RETURNING id"),
        {"company_id": company_id, "group_name": group_name, "address": address, "former_address": former_address, "property_type": property_type},
    )
    return int(property_id), True


async def upsert_unit(db: AsyncSession, property_id: int, unit_label: str | None, is_basement: bool) -> tuple[int, bool]:
    existing = await db.scalar(
        text("SELECT id FROM rental_units WHERE property_id=:property_id AND unit_label IS NOT DISTINCT FROM :unit_label"),
        {"property_id": property_id, "unit_label": unit_label},
    )
    if existing is not None:
        await db.execute(text("UPDATE rental_units SET is_basement=:is_basement, updated_at=now() WHERE id=:id"), {"id": existing, "is_basement": is_basement})
        return int(existing), False
    unit_id = await db.scalar(
        text("INSERT INTO rental_units (property_id, unit_label, is_basement) VALUES (:property_id, :unit_label, :is_basement) RETURNING id"),
        {"property_id": property_id, "unit_label": unit_label, "is_basement": is_basement},
    )
    return int(unit_id), True


async def upsert_tenant(db: AsyncSession, tenant: dict[str, Any]) -> tuple[int, bool]:
    tenant_id = await db.scalar(text("SELECT id FROM rental_tenants WHERE lower(full_name)=lower(:name)"), {"name": tenant["full_name"]})
    values = {"name": tenant["full_name"].strip(), "phone": tenant.get("phone"), "email": tenant.get("email"), "secondary_email": tenant.get("secondary_email")}
    if tenant_id is not None:
        await db.execute(text("UPDATE rental_tenants SET phone=:phone, email=:email, secondary_email=:secondary_email, updated_at=now() WHERE id=:id"), {**values, "id": tenant_id})
        return int(tenant_id), False
    tenant_id = await db.scalar(text("INSERT INTO rental_tenants (full_name, phone, email, secondary_email) VALUES (:name, :phone, :email, :secondary_email) RETURNING id"), values)
    return int(tenant_id), True


async def upsert_lease(db: AsyncSession, unit_id: int, row: tuple[Any, ...], seed: dict[str, Any]) -> tuple[int, bool, bool, str]:
    rent = decimal_or_none(row[RENT_COLUMNS["rent"]])
    if rent is None:
        raise ValueError(f"Rent is missing for rent-roll row {row[0]}")
    deposit = decimal_or_none(row[RENT_COLUMNS["deposit"]])
    discount, discount_raw = parse_discount(row[RENT_COLUMNS["discount"]])
    lease_start, start_raw = parse_date(row[RENT_COLUMNS["lease_start"]])
    lease_end, end_raw = parse_date(row[RENT_COLUMNS["lease_end"]])
    parse_flag = bool(start_raw or end_raw)
    notes: list[str] = []
    append_notes(notes, row[RENT_COLUMNS["notes"]])
    append_notes(notes, seed.get("notes"))
    append_notes(notes, seed.get("lease_notes_append"))
    for tenant in seed["tenants"]:
        if tenant.get("notes"):
            notes.append(f"[tenant check: {tenant['full_name']}] {tenant['notes']}")
    if start_raw:
        notes.append(f"[unparsed date] lease_start={start_raw}")
    if end_raw:
        notes.append(f"[unparsed date] lease_end={end_raw}")
    lease_status = "expired" if lease_end is not None and lease_end < date.today() else "active"
    params = {
        "unit_id": unit_id, "rent": rent, "discount": discount, "discount_raw": discount_raw,
        "deposit": deposit, "lease_start": lease_start, "lease_end": lease_end,
        "parse_flag": parse_flag, "status": lease_status, "notes": "\n".join(notes) or None,
    }
    lease_id = await db.scalar(
        text("SELECT id FROM rental_leases WHERE unit_id=:unit_id AND lease_start IS NOT DISTINCT FROM :lease_start AND rent=:rent"),
        params,
    )
    created = lease_id is None
    if created:
        lease_id = await db.scalar(text("INSERT INTO rental_leases (unit_id, rent, rent_discount_amount, rent_discount_raw, deposit, lease_start, lease_end, date_parse_flag, status, lease_notes) VALUES (:unit_id, :rent, :discount, :discount_raw, :deposit, :lease_start, :lease_end, :parse_flag, :status, :notes) RETURNING id"), params)
    else:
        await db.execute(text("UPDATE rental_leases SET rent_discount_amount=:discount, rent_discount_raw=:discount_raw, deposit=:deposit, lease_end=:lease_end, date_parse_flag=:parse_flag, status=:status, lease_notes=:notes, updated_at=now() WHERE id=:id"), {**params, "id": lease_id})
    return int(lease_id), created, parse_flag, lease_status


async def seed(args: argparse.Namespace) -> None:
    rent_book = load_workbook(args.rent_roll, read_only=True, data_only=True)["Sheet1"]
    inspection_book = load_workbook(args.inspections, read_only=True, data_only=True)["Sheet1"]
    tenant_seed = json.loads(Path(args.tenants_json).read_text())["leases"]
    tenants_by_row = {int(entry["rent_roll_row_num"]): entry for entry in tenant_seed}
    rent_rows = {int(row[0]): row for row in rent_book.iter_rows(min_row=2, values_only=True) if isinstance(row[0], (int, float))}
    mapping_by_rent = {int(mapping[5]): mapping for mapping in UNIT_MAPPINGS if mapping[5] is not None}
    mapping_by_inspection = {str(mapping[6]).strip(): mapping for mapping in UNIT_MAPPINGS}
    unmatched_rent = sorted(set(rent_rows) - set(mapping_by_rent))
    missing_tenants = sorted(set(mapping_by_rent) - set(tenants_by_row))
    if unmatched_rent or missing_tenants:
        raise ValueError(f"Mapping validation failed: unmatched rent rows={unmatched_rent}; missing tenant seed rows={missing_tenants}")

    counts = {key: 0 for key in ("properties", "units", "tenants", "leases", "inspections")}
    date_flags: list[int] = []
    expired = 0
    unit_by_rent: dict[int, int] = {}
    unit_by_inspection: dict[str, int] = {}
    async with AsyncSessionLocal() as db:
        try:
            company_id, _ = await get_or_create_company(db)
            property_ids: dict[str, int] = {}
            for mapping in UNIT_MAPPINGS:
                address = mapping[1]
                if address not in property_ids:
                    property_id, created = await upsert_property(db, company_id, mapping)
                    property_ids[address] = property_id
                    counts["properties"] += int(created)
                unit_id, created = await upsert_unit(db, property_ids[address], mapping[3], bool(mapping[4]))
                counts["units"] += int(created)
                if mapping[5] is not None:
                    unit_by_rent[int(mapping[5])] = unit_id
                unit_by_inspection[str(mapping[6]).strip()] = unit_id

            for row_number, row in rent_rows.items():
                unit_id = unit_by_rent[row_number]
                water_amount, water_raw = parse_water(row[RENT_COLUMNS["water"]])
                await db.execute(text("UPDATE rental_units SET water_credit_amount=:amount, water_deal_raw=:raw, status='occupied', updated_at=now() WHERE id=:id"), {"id": unit_id, "amount": water_amount, "raw": water_raw})
                lease_seed = tenants_by_row[row_number]
                lease_id, created, parse_flag, lease_status = await upsert_lease(db, unit_id, row, lease_seed)
                counts["leases"] += int(created)
                if parse_flag:
                    date_flags.append(row_number)
                expired += int(lease_status == "expired")
                for position, tenant in enumerate(lease_seed["tenants"]):
                    tenant_id, tenant_created = await upsert_tenant(db, tenant)
                    counts["tenants"] += int(tenant_created)
                    await db.execute(text("INSERT INTO rental_lease_tenants (lease_id, tenant_id, is_primary_contact) VALUES (:lease_id, :tenant_id, :primary) ON CONFLICT (lease_id, tenant_id) DO UPDATE SET is_primary_contact=EXCLUDED.is_primary_contact"), {"lease_id": lease_id, "tenant_id": tenant_id, "primary": position == 0})

            today = date.today()
            for row in inspection_book.iter_rows(min_row=3, values_only=True):
                label = "" if row[0] is None else str(row[0]).strip()
                if label not in mapping_by_inspection or not any(value is not None for value in row[1:]):
                    continue
                unit_id = unit_by_inspection[label]
                front_score, front_notes = score_and_notes(row[1])
                back_score, back_notes = score_and_notes(row[2])
                general_notes = None if row[4] is None else str(row[4]).strip()
                occupancy = "vacant" if general_notes and "vacant" in general_notes.casefold() else None
                result = await db.execute(text("INSERT INTO rental_inspections (unit_id, inspection_type, inspection_date, inspector_name, front_yard_score, front_yard_notes, back_yard_score, back_yard_notes, building_condition, occupancy_flag, general_notes, status) VALUES (:unit_id, 'exterior', :today, 'Nicholas', :front_score, :front_notes, :back_score, :back_notes, :building, :occupancy, :general_notes, 'submitted') ON CONFLICT (unit_id, inspection_date, inspection_type) DO UPDATE SET inspector_name=EXCLUDED.inspector_name, front_yard_score=EXCLUDED.front_yard_score, front_yard_notes=EXCLUDED.front_yard_notes, back_yard_score=EXCLUDED.back_yard_score, back_yard_notes=EXCLUDED.back_yard_notes, building_condition=EXCLUDED.building_condition, occupancy_flag=EXCLUDED.occupancy_flag, general_notes=EXCLUDED.general_notes, updated_at=now() RETURNING (xmax = 0) AS inserted"), {"unit_id": unit_id, "today": today, "front_score": front_score, "front_notes": front_notes, "back_score": back_score, "back_notes": back_notes, "building": row[3], "occupancy": occupancy, "general_notes": general_notes})
                counts["inspections"] += int(bool(result.scalar_one()))
            await db.commit()
        except Exception:
            await db.rollback()
            raise

    print("PRIVI rentals import complete")
    print("Created:", ", ".join(f"{key}={value}" for key, value in counts.items()))
    print(f"Source totals: properties={len({mapping[1] for mapping in UNIT_MAPPINGS})}, units={len(UNIT_MAPPINGS)}, leases={len(rent_rows)}")
    print(f"Date parse flags (rent-roll row numbers): {date_flags or 'none'}")
    print(f"Leases auto-marked expired: {expired} — review for month-to-month status")
    print("UNCONFIRMED: 59 Guay Avenue Units 3 and 4 marked basement from inspection labels — check with Nicholas")
    print(f"Unmatched rent-roll rows: {unmatched_rent or 'none'}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Idempotently seed PRIVI rental data")
    parser.add_argument("--rent-roll", type=Path, required=True)
    parser.add_argument("--inspections", type=Path, required=True)
    parser.add_argument("--tenants-json", type=Path, required=True)
    return parser.parse_args()


if __name__ == "__main__":
    asyncio.run(seed(parse_args()))
